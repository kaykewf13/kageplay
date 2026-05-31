#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KagePlay - scripts/build.py
Le a planilha-mestre (.xlsx) e gera os arquivos do app/player:
    data/catalog.json
    data/categories.json
    data/playlist.m3u
Tambem emite um relatorio de inconsistencias.

Uso:
    python scripts/build.py --xlsx kageplay_planilha_mestre.xlsx --out data
Requisitos:
    pip install openpyxl
"""
import argparse, json, sys, re
from pathlib import Path
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRO: instale openpyxl -> pip install openpyxl"); sys.exit(1)


def norm(v):
    return ("" if v is None else str(v)).strip()


def sim(v):
    return norm(v).lower() in ("sim", "yes", "true", "1")


def slugify(s):
    s = norm(s).lower()
    s = re.sub(r"[áàâã]", "a", s); s = re.sub(r"[éê]", "e", s)
    s = re.sub(r"[íï]", "i", s);   s = re.sub(r"[óôõ]", "o", s)
    s = re.sub(r"[úü]", "u", s);   s = re.sub(r"ç", "c", s)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def read_sheet(wb, name):
    """Le uma aba como lista de dicts usando a 1a linha como cabecalho."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = [norm(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r): continue
        out.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})
    return out


def elegivel(ep):
    return sim(ep.get("gratuito_autorizado")) and not sim(ep.get("drm_paywall")) and sim(ep.get("publicar"))


def build(xlsx_path, out_dir):
    wb = load_workbook(xlsx_path, data_only=True)
    animes_raw = read_sheet(wb, "01_Animes")
    eps_raw    = read_sheet(wb, "02_Episodios")
    cats_raw   = read_sheet(wb, "03_Categorias")

    relatorio = {"erros": [], "alertas": []}

    # ---- indexar episodios por anime_id ----
    eps_by_anime = {}
    anime_ids = {norm(a.get("anime_id")) for a in animes_raw}
    for e in eps_raw:
        aid = norm(e.get("anime_id"))
        if not aid:
            relatorio["erros"].append("Episodio sem anime_id."); continue
        if aid not in anime_ids:
            relatorio["erros"].append(f"Episodio orfao: anime_id '{aid}' nao existe em 01_Animes.")
        eps_by_anime.setdefault(aid, []).append(e)

    # ---- montar catalogo ----
    animes_out = []
    for a in animes_raw:
        aid = norm(a.get("anime_id"))
        if not aid:
            relatorio["erros"].append("Anime sem anime_id (ignorado)."); continue
        if not sim(a.get("publicar")):
            continue  # nao publicado nao entra no catalogo

        eps_out = []
        for e in sorted(eps_by_anime.get(aid, []), key=lambda x: int(norm(x.get("episodio_num")) or 0)):
            if norm(e.get("status_link")) and norm(e.get("status_link")) != "Ativo":
                relatorio["alertas"].append(
                    f"[{aid} ep {norm(e.get('episodio_num'))}] link {norm(e.get('status_link'))} - alerta operacional na planilha.")
            eps_out.append({
                "episodio_num": int(norm(e.get("episodio_num")) or 0),
                "titulo_episodio": norm(e.get("titulo_episodio")),
                "url_video": norm(e.get("url_video")),
                "fonte_url": norm(e.get("fonte_url")) or norm(e.get("url_fonte")),
                "tipo_player": norm(e.get("tipo_player")) or "externo",
                "modo_reproducao": norm(e.get("modo_reproducao")) or "interno",
                "gratuito_autorizado": "Sim" if sim(e.get("gratuito_autorizado")) else "Nao",
                "drm_paywall": "Sim" if sim(e.get("drm_paywall")) else "Nao",
                "publicar": "Sim" if sim(e.get("publicar")) else "Nao",
                "status_link": norm(e.get("status_link")) or "Ativo",
                "adulto_18": sim(e.get("adulto_18")),
            })

        if not eps_out:
            relatorio["alertas"].append(f"Anime '{aid}' publicado mas sem episodios.")

        sec = norm(a.get("categorias_secundarias"))
        animes_out.append({
            "anime_id": aid,
            "titulo": norm(a.get("titulo")),
            "slug": norm(a.get("slug")) or slugify(a.get("titulo")),
            "categoria_principal": norm(a.get("categoria_principal")),
            "categorias_secundarias": [c.strip() for c in sec.split(",") if c.strip()],
            "adulto_18": sim(a.get("adulto_18")),
            "status": norm(a.get("status")) or "Em revisao",
            "publicar": "Sim",
            "descricao": norm(a.get("descricao")),
            "capa_url": norm(a.get("capa_url")),
            "banner_url": norm(a.get("banner_url")),
            "fonte_principal": norm(a.get("fonte_principal")),
            "episodios": eps_out,
        })

    catalog = {"versao": "2.1", "atualizado_em": str(date.today()), "animes": animes_out}

    # ---- categorias ----
    cats_out = []
    for c in cats_raw:
        nome = norm(c.get("categoria"))
        if not nome: continue
        cats_out.append({
            "grupo": norm(c.get("grupo")),
            "categoria": nome,
            "slug": norm(c.get("slug")) or slugify(nome),
            "adulto_18": sim(c.get("adulto_18")),
            "mostrar_menu": sim(c.get("mostrar_menu")),
            "template_canva": norm(c.get("template_canva")) or "categoria_padrao",
        })
    categories = {"versao": "2.1", "atualizado_em": str(date.today()), "categorias": cats_out}

    # ---- playlist.m3u (apenas elegiveis e mp4/hls/webm) ----
    m3u = ["#EXTM3U"]
    for a in animes_out:
        for e in a["episodios"]:
            if elegivel(e) and e["tipo_player"] in ("mp4", "webm", "hls"):
                grp = "18+" if a["adulto_18"] else a["categoria_principal"]
                m3u.append(f'#EXTINF:-1 group-title="{grp}",{a["titulo"]} - Ep {e["episodio_num"]}')
                m3u.append(e["url_video"])

    # ---- gravar ----
    out = Path(out_dir); out.mkdir(parents=True, exist_ok=True)
    (out / "catalog.json").write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "categories.json").write_text(json.dumps(categories, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "playlist.m3u").write_text("\n".join(m3u) + "\n", encoding="utf-8")

    # ---- resumo ----
    total_eps = sum(len(a["episodios"]) for a in animes_out)
    eleg = sum(1 for a in animes_out for e in a["episodios"] if elegivel(e))
    print("=" * 56)
    print(f" KagePlay build OK")
    print(f"  Animes publicados : {len(animes_out)}")
    print(f"  Episodios         : {total_eps}")
    print(f"  Elegiveis player  : {eleg}")
    print(f"  Categorias        : {len(cats_out)}")
    print(f"  Erros             : {len(relatorio['erros'])}")
    print(f"  Alertas (planilha): {len(relatorio['alertas'])}")
    print("=" * 56)
    for x in relatorio["erros"]:   print("  [ERRO]   ", x)
    for x in relatorio["alertas"]: print("  [ALERTA] ", x)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Gera JSON/M3U do KagePlay a partir da planilha.")
    ap.add_argument("--xlsx", required=True, help="Caminho da planilha-mestre .xlsx")
    ap.add_argument("--out", default="data", help="Pasta de saida (default: data)")
    args = ap.parse_args()
    build(args.xlsx, args.out)
